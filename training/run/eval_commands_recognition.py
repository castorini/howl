from .args import ArgumentParserBuilder, opt
import random
import os
import subprocess
import numpy as np

from datetime import datetime
from openpyxl import Workbook

"""
This script uses every GPU on the machine to train every models in this repo for detecting gsc commands and compare their performance

target keywords include “yes,” “no,” “up,” “down,” “left,” “right,” “on,” “off,”, “stop,” “go,” unknown, or silence.

sample command:
python -m training.run.eval_commands_recognition --num_iterations x --dataset_path < path_to_gsc_datasets >
"""

def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--num_iterations',
                        type=int,
                        default=1,
                        help='number of experiments to run'),
                    opt('--dataset_path',
                        type=str,
                        default="/data/kws/gsc/v1"))

    args = apb.parser.parse_args()

    # preparing xlsx objects
    wb = Workbook()
    now = datetime.now()
    dev_sheet = wb.create_sheet('dev', 0)
    test_sheet = wb.create_sheet('test', 1)

    model_types = ['res8', 'las', 'lstm', 'mobilenet']
    col_index_map = {
        'res8': 'B',
        'las': 'C',
        'lstm': 'D',
        'mobilenet': 'E'
    }
    dev_sheet['A2'] = "mean"
    dev_sheet['A3'] = "std"
    dev_sheet['A4'] = "p90"
    dev_sheet['A5'] = "p95"
    dev_sheet['A6'] = "p99"
    
    dev_sheet['B1'] = model_types[0]
    dev_sheet['C1'] = model_types[1]
    dev_sheet['D1'] = model_types[2]
    dev_sheet['E1'] = model_types[3]

    test_sheet['A2'] = "mean"
    test_sheet['A3'] = "std"
    test_sheet['A4'] = "p90"
    test_sheet['A5'] = "p95"
    test_sheet['A6'] = "p99"

    test_sheet['B1'] = model_types[0]
    test_sheet['C1'] = model_types[1]
    test_sheet['D1'] = model_types[2]
    test_sheet['E1'] = model_types[3]

    os.system("mkdir -p exp_results")
    dt_string = now.strftime("%b-%d-%H-%M")
    file_name = "exp_results/" + dt_string + ".xlsx"
    print("report will be generated at ", file_name)

    # training settings
    os.environ["DATASET_PATH"] = args.dataset_path
    os.environ["BATCH_SIZE"] = "64"
    os.environ["MAX_WINDOW_SIZE_SECONDS"] = "1"
    os.environ["USE_NOISE_DATASET"] = "False"
    os.environ["INFERENCE_SEQUENCE"] = "[0]"
    os.environ["WEIGHT_DECAY"] = "0.00001"
    os.environ["NUM_EPOCHS"] = "20"
    os.environ["LR_DECAY"] = "0.8"
    os.environ["NUM_MELS"] = "40"
    os.environ["VOCAB"] = '["yes","no","up","down","left","right","on","off","stop","go"]'

    results = {
        'res8': [[],[]],
        'las': [[],[]],
        'lstm': [[],[]],
        'mobilenet': [[],[]],
    }

    for i in range(args.num_iterations):
        print("\titeration: ", i, "/", args.num_iterations)
        os.environ["SEED"] = str(random.randint(1,1000000))

        row_index = str(i + 8)
        dev_sheet['A'+row_index] = os.environ["SEED"]
        test_sheet['A'+row_index] = os.environ["SEED"]

        for model_type in model_types:
            print("\tmodel: ", model_type, " - ", datetime.now().strftime("%H:%M"), flush=True)

            if model_type == 'res8':
                os.environ["LEARNING_RATE"] = "0.01"
            else:
                os.environ["LEARNING_RATE"] = "0.001"

            workspace_path = os.getcwd() + "/workspaces/" + model_type + "/" + str(i)
            log_path =  workspace_path + "/exp.log"

            os.system("mkdir -p " + workspace_path)
            log_path =  workspace_path + "/exp.log"
            exp_execution = os.system("touch " + log_path)
            exp_execution = os.system("python -m training.run.pretrain_gsc --model " + model_type + " --workspace " + workspace_path + " 2>&1 | tee " + log_path)

            # start training
            raw_log = subprocess.check_output(['cat', log_path]).decode("utf-8") 

            # process the results
            logs = raw_log.split('\n')
            dev_acc, test_acc = logs[-3:-1]
            dev_acc = float(dev_acc.split(' ')[2])
            test_acc = float(test_acc.split(' ')[2])

            index = col_index_map[model_type] + row_index
            dev_sheet[index] = dev_acc
            test_sheet[index] = test_acc
            results[model_type][0].append(dev_acc)
            results[model_type][1].append(test_acc)

            dev_metrics = np.array(results[model_type][0])
            dev_sheet[col_index_map[model_type] + '2'] = str(dev_metrics.mean())
            dev_sheet[col_index_map[model_type] + '3'] = str(dev_metrics.std())
            dev_sheet[col_index_map[model_type] + '4'] = str(np.percentile(dev_metrics, 90))
            dev_sheet[col_index_map[model_type] + '5'] = str(np.percentile(dev_metrics, 95))
            dev_sheet[col_index_map[model_type] + '6'] = str(np.percentile(dev_metrics, 99))

            test_metrics = np.array(results[model_type][1])
            test_sheet[col_index_map[model_type] + '2'] = str(test_metrics.mean())
            test_sheet[col_index_map[model_type] + '3'] = str(test_metrics.std())
            test_sheet[col_index_map[model_type] + '4'] = str(np.percentile(test_metrics, 90))
            test_sheet[col_index_map[model_type] + '5'] = str(np.percentile(test_metrics, 95))
            test_sheet[col_index_map[model_type] + '6'] = str(np.percentile(test_metrics, 99))

        wb.save(file_name)


if __name__ == '__main__':
    main()
