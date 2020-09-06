from .args import ArgumentParserBuilder, opt
import random
import os
import subprocess
import numpy as np

from datetime import datetime
from openpyxl import Workbook, load_workbook


def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--n',
                        type=int,
                        default=1,
                        help='number of experiments to run'),
                    opt('--dataset_path',
                        type=str,
                        default="/data/kws/gsc/v1"))

    args = apb.parser.parse_args()

    # Create new workbook
    wb = Workbook()
    now = datetime.now()
    dev_sheet = wb.create_sheet('dev', 0)
    test_sheet = wb.create_sheet('test', 1)

    model_types = ['res8', 'las', 'lstm', 'mobilenet']
    col_map = {
        'res8': 'B',
        'las': 'C',
        'lstm': 'D',
        'mobilenet': 'E'
    }
    dev_sheet['A2'] = "mean"
    dev_sheet['A3'] = "std"
    dev_sheet['A4'] = "p90"
    dev_sheet['A5'] = "p95"
    dev_sheet['A6'] = "p99"


    dev_sheet['B1'] = model_types[0]
    dev_sheet['C1'] = model_types[1]
    dev_sheet['D1'] = model_types[2]
    dev_sheet['E1'] = model_types[3]

    test_sheet['B1'] = model_types[0]
    test_sheet['C1'] = model_types[1]
    test_sheet['D1'] = model_types[2]
    test_sheet['E1'] = model_types[3]

    os.system("mkdir -p exp_results")
    dt_string = now.strftime("%b-%d-%H-%M")
    file_name = "exp_results/" + dt_string + ".xlsx"
    print("exp_result stored at ", file_name)


    os.environ["DATASET_PATH"] = args.dataset_path
    os.environ["BATCH_SIZE"] = "64"
    os.environ["MAX_WINDOW_SIZE_SECONDS"] = "1"
    os.environ["USE_NOISE_DATASET"] = "False"
    os.environ["INFERENCE_SEQUENCE"] = "[0]"

    os.environ["WEIGHT_DECAY"] = "0.00001"
    os.environ["NUM_EPOCHS"] = "20"
    os.environ["LR_DECAY"] = "0.8"
    os.environ["NUM_MELS"] = "40"

    vocabs = [ name for name in os.listdir(args.dataset_path) if os.path.isdir(os.path.join(args.dataset_path, name)) ]
    vocabs.remove('_background_noise_')

    results = {
        'res8': [[],[]],
        'las': [[],[]],
        'lstm': [[],[]],
        'mobilenet': [[],[]],
    }

    for i in range(args.n):
        print("\titeration: ", i )
        # print("\tvocabs: ", selected_vocab)
        # os.environ["VOCAB"] = str(selected_vocab).replace(" ", "").replace("\'", "\"")
        os.environ["VOCAB"] = '["yes","no","up","down","left","right","on","off","stop","go"]'
        os.environ["SEED"] = str(random.randint(1,1000000))

        row_index = str(i + 8)
        dev_sheet['A'+row_index] = str(i)
        test_sheet['A'+row_index] = str(i)

        for model_type in model_types:
            print("\tmodel: ", model_type, " - ", datetime.now().strftime("%H-%M"), flush=True)

            if model_type == 'res8':
                os.environ["LEARNING_RATE"] = "0.01"
            else:
                os.environ["LEARNING_RATE"] = "0.001"

            workspace_path = os.getcwd() + "/workspaces/" + model_type + "/" + str(i)
            log_path =  workspace_path + "/exp.log"

            os.system("mkdir -p " + workspace_path)
            log_path =  workspace_path + "/exp.log"
            exp_execution = os.system("touch " + log_path)
            exp_execution = os.system("python -m howl.run.pretrain_gsc --model " + model_type + " --workspace " + workspace_path + " 2>&1 | tee " + log_path)

            raw_log = subprocess.check_output(['cat', log_path]).decode("utf-8") 
            logs = raw_log.split('\n')

            dev_acc, test_acc = logs[-3:-1]
            dev_acc = float(dev_acc.split(' ')[2])
            test_acc = float(test_acc.split(' ')[2])

            index = col_map[model_type] + row_index

            dev_sheet[index] = dev_acc
            test_sheet[index] = test_acc

            results[model_type][0].append(dev_acc)
            results[model_type][1].append(test_acc)

            dev_metrics = np.array(results[model_type][0])
            dev_sheet[col_map[model_type] + '2'] = str(dev_metrics.mean())
            dev_sheet[col_map[model_type] + '3'] = str(dev_metrics.std())
            dev_sheet[col_map[model_type] + '4'] = str(np.percentile(dev_metrics, 90))
            dev_sheet[col_map[model_type] + '5'] = str(np.percentile(dev_metrics, 95))
            dev_sheet[col_map[model_type] + '6'] = str(np.percentile(dev_metrics, 99))

            test_metrics = np.array(results[model_type][1])
            test_sheet[col_map[model_type] + '2'] = str(test_metrics.mean())
            test_sheet[col_map[model_type] + '3'] = str(test_metrics.std())
            test_sheet[col_map[model_type] + '4'] = str(np.percentile(test_metrics, 90))
            test_sheet[col_map[model_type] + '5'] = str(np.percentile(test_metrics, 95))
            test_sheet[col_map[model_type] + '6'] = str(np.percentile(test_metrics, 99))

        wb.save(file_name)




if __name__ == '__main__':
    main()
