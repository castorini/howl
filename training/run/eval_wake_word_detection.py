import os
import random
import subprocess
import sys
import time
from datetime import datetime
from os import path

import numpy as np
import torch
from openpyxl import Workbook

from .args import ArgumentParserBuilder, opt

"""
This script uses every GPU on the machine to train multiple res8 models for hey_firefox and hey_snips,
evaluates them with different threashold, and generate performance reports

Two different performance reports are generated, one with the clean audio and one with audios with noise

sample command:
python -m training.run.eval_wake_word_detection --num_models 10 --hop_size 0.05 --exp_type hey_firefox --dataset_path "x" --noiseset_path "y"

train 10 models for hey firefox with random seeds using datasets x and y
then evaluate them for thresholds ranging from 0 to 1 in increments of 0.05
"""


def is_job_running(grep_command, count_command):
    out = subprocess.check_output('ps aux | grep ' + grep_command, shell=True)
    num_proc = out.decode('utf-8').count(count_command)
    return num_proc > 0


def run_batch_commands(commands, envs, grep_command='training.run.train', count_command='python -m training.run.train'):
    """
    run given set of commands with the corresponding environments
    check the status of each process regularly and schedule the next job whenever GPU is availble
    """

    num_gpu = torch.cuda.device_count()
    print('availble number of GPU is', num_gpu)

    check_up_delay = 600  # check every 10 mins
    num_running_jobs = 0
    sleep_counter = 0

    for (command, env) in zip(commands, envs):
        for env_key, env_val in env.items():
            os.environ[env_key] = env_val

        if num_running_jobs == num_gpu:
            sleep_counter = 0
            while is_job_running(grep_command, count_command):
                sleep_counter += 1
                print('some jobs are running; waiting for {} mins'.format(sleep_counter * check_up_delay / 60))
                sys.stdout.flush()
                sys.stderr.flush()
                time.sleep(check_up_delay)
            num_running_jobs = 0

        new_env = os.environ.copy()
        new_env['CUDA_VISIBLE_DEVICES'] = str(num_running_jobs)
        proc = subprocess.Popen(command.split(),
                                preexec_fn=os.setpgrp,
                                env=new_env)

        print('process {} - '.format(proc.pid), command, new_env, flush=True)
        time.sleep(60)  # add some delay between each job scheduling
        num_running_jobs += 1

    sleep_counter = 0
    while is_job_running(grep_command, count_command):
        sleep_counter += 1
        print('some jobs are running; waiting for {} mins'.format(sleep_counter * check_up_delay / 60))
        time.sleep(check_up_delay)


def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--num_models',
                        type=int,
                        default=1,
                        help='number of models to train and evaluate'),
                    opt('--hop_size',
                        type=float,
                        default=0.05,
                        help='hop size for threshold'),
                    opt('--dataset_path',
                        type=str,
                        default='/data/speaker-id-split-medium'),
                    opt('--exp_type',
                        type=str, choices=['hey_firefox', 'hey_snips'], default='hey_firefox'),
                    opt('--seed',
                        type=int, default=0),
                    opt('--noiseset_path',
                        type=str,
                        default='/data/MS-SNSD'))

    args = apb.parser.parse_args()

    random.seed(args.seed)

    # Preapring xlsx objects
    clean_wb = Workbook()
    noisy_wb = Workbook()
    now = datetime.now()
    clean_sheets = {}
    noisy_sheets = {}

    col_mapping = {
        'Dev positive': 'B',
        'Dev noisy positive': 'B',
        'Dev negative': 'H',
        'Dev noisy negative': 'H',
        'Test positive': 'N',
        'Test noisy positive': 'N',
        'Test negative': 'T',
        'Test noisy negative': 'T'
    }

    # total sample counts, these number can be found when the datasets are loaded for training
    if args.exp_type == "hey_firefox":
        total_counts = {
            'Dev positive': '76',
            'Dev negative': '2531',
            'Test positive': '54',
            'Test negative': '2504'
        }
    elif args.exp_type == "hey_snips":
        total_counts = {
            'Dev positive': '2484',
            'Dev negative': '13598',
            'Test positive': '2529',
            'Test negative': '13943'
        }

    metrics = ['threshold', 'tp', 'tn', 'fp', 'fn']
    clean_col_names = ['Dev positive', 'Dev negative', 'Test positive', 'Test negative']

    def round_and_convert_to_str(n):
        return str(round(n, 2))

    raw_thresholds = np.arange(0, 1.000001, args.hop_size)
    thresholds = list(map(round_and_convert_to_str, raw_thresholds))

    # compute metrics for the overall experiments
    clean_cols_aggregated = {}
    noisy_cols_aggregated = {}

    def compute_aggregated_metrics(sheet, col_idx, results):
        sheet[col_idx + '3'] = str(results.mean())
        sheet[col_idx + '4'] = str(results.std())
        sheet[col_idx + '5'] = str(np.percentile(results, 90))
        sheet[col_idx + '6'] = str(np.percentile(results, 95))
        sheet[col_idx + '7'] = str(np.percentile(results, 99))
        sheet[col_idx + '8'] = str(results.sum())

    def get_cell_idx(char, ind):
        return chr(ord(char) + ind)

    def prepare_report(sheet):
        sheet['A3'] = 'mean'
        sheet['A4'] = 'std'
        sheet['A5'] = 'p90'
        sheet['A6'] = 'p95'
        sheet['A7'] = 'p99'
        sheet['A8'] = 'sum'

        for col in clean_col_names:
            cell_idx = col_mapping[col] + '1'
            sheet[cell_idx] = col

            cell_idx = get_cell_idx(col_mapping[col], 1) + '1'
            sheet[cell_idx] = total_counts[col]

            for i, metric in enumerate(metrics):
                cell_idx = get_cell_idx(col_mapping[col], i) + '2'
                sheet[cell_idx] = metric

    # reports are generated for each threshold and each gets a separate sheet
    for idx, threshold in enumerate(thresholds):
        clean_sheets[threshold] = clean_wb.create_sheet(threshold, idx)
        prepare_report(clean_sheets[threshold])
        noisy_sheets[threshold] = noisy_wb.create_sheet(threshold, idx)
        prepare_report(noisy_sheets[threshold])

        for col_name in clean_col_names:
            for metric_idx, metric in enumerate(metrics):
                target_metric = threshold + '_' + get_cell_idx(col_mapping[col_name], metric_idx)
                clean_cols_aggregated[target_metric] = []
                noisy_cols_aggregated[target_metric] = []

    # reports are generated at exp_results
    os.system('mkdir -p exp_results')
    dt_string = now.strftime('%b-%d-%H-%M')

    clean_file_name = 'exp_results/'+args.exp_type+'_clean_' + dt_string + '.xlsx'
    clean_wb.save(clean_file_name)
    print('\treport for clean setting is generated at ', clean_file_name)

    noisy_file_name = 'exp_results/'+args.exp_type+'_noisy_' + dt_string + '.xlsx'
    noisy_wb.save(noisy_file_name)
    print('\treport for noisy setting is generated at ', noisy_file_name)

    # Training settings
    os.environ['DATASET_PATH'] = args.dataset_path
    os.environ['WEIGHT_DECAY'] = '0.00001'
    os.environ['LEARNING_RATE'] = '0.01'
    os.environ['LR_DECAY'] = '0.98'
    os.environ['BATCH_SIZE'] = '16'
    os.environ['MAX_WINDOW_SIZE_SECONDS'] = '0.5'
    os.environ['USE_NOISE_DATASET'] = 'True'
    os.environ['NUM_MELS'] = '40'
    os.environ['NOISE_DATASET_PATH'] = args.noiseset_path

    if args.exp_type == "hey_firefox":
        os.environ['NUM_EPOCHS'] = '300'
        os.environ['VOCAB'] = '["hey","fire","fox"]'
        os.environ['INFERENCE_SEQUENCE'] = '[0,1,2]'
    elif args.exp_type == "hey_snips":
        os.environ['NUM_EPOCHS'] = '100'
        os.environ['VOCAB'] = '["hey","snips"]'
        os.environ['INFERENCE_SEQUENCE'] = '[0,1]'

    seeds = []

    print('-- training ', args.num_models, ' models --')
    training_commands = []
    training_envs = []

    def get_workspace_path(exp_type, seed):
        return os.getcwd() + '/workspaces/exp_' + exp_type + '_res8/' + str(seed)

    # generate commands to run along with the environments
    for i in range(args.num_models):
        seed = str(random.randint(1, 1000000))
        seeds.append(seed)
        env = {}
        env['SEED'] = seed
        workspace_path = get_workspace_path(args.exp_type, seed)
        os.system('mkdir -p ' + workspace_path)
        command = 'python -m training.run.train --model res8 --workspace ' + workspace_path + '  -i ' + args.dataset_path
        training_commands.append(command)
        training_envs.append(env)

    print('seeds for each model: ', seeds)

    run_batch_commands(training_commands, training_envs)

    print('-- evaluating each models --')
    eval_commands = []
    eval_envs = []

    for seed in seeds:
        for threshold_idx, threshold in enumerate(thresholds):
            env = {}
            env['SEED'] = seed
            env['INFERENCE_THRESHOLD'] = threshold
            workspace_path = get_workspace_path(args.exp_type, seed)
            command = 'python -m training.run.train --eval --model res8 --workspace ' + workspace_path + '  -i ' + args.dataset_path
            result_path = workspace_path + '/' + threshold + '_results.csv'

            # if evaluation is done previously, we skip it
            if not path.exists(result_path):
                eval_commands.append(command)
                eval_envs.append(env)

    run_batch_commands(eval_commands, eval_envs)

    print('-- generating reports --')

    for seed_idx, seed in enumerate(seeds):
        for threshold_idx, threshold in enumerate(thresholds):
            clean_sheet = clean_sheets[threshold]
            noisy_sheet = noisy_sheets[threshold]

            row = str(seed_idx + 10)
            clean_sheet['A'+row] = seed
            noisy_sheet['A'+row] = seed

            workspace_path = get_workspace_path(args.exp_type, seed)
            result_path = workspace_path + '/' + threshold + '_results.csv'
            raw_result = subprocess.check_output(['tail', '-n', '8', result_path]).decode('utf-8')
            results = raw_result.split('\n')

            # parse and update the report
            for result in results:
                if len(result) == 0:
                    break
                vals = result.split(',')
                key = vals[0]
                start_col = col_mapping[key]

                if 'noisy' in key:
                    for metric_idx, metric in enumerate(vals[1:]):
                        col_idx = get_cell_idx(start_col, metric_idx)
                        cell_ind = col_idx + str(row)
                        noisy_sheet[cell_ind] = metric

                        target_metric = threshold + '_' + col_idx
                        noisy_cols_aggregated[target_metric].append(float(metric))

                        compute_aggregated_metrics(noisy_sheet, col_idx, np.array(noisy_cols_aggregated[target_metric]))
                else:
                    for metric_idx, metric in enumerate(vals[1:]):
                        col_idx = get_cell_idx(start_col, metric_idx)
                        cell_ind = col_idx + str(row)
                        clean_sheet[cell_ind] = metric

                        target_metric = threshold + '_' + col_idx
                        clean_cols_aggregated[target_metric].append(float(metric))

                        compute_aggregated_metrics(clean_sheet, col_idx, np.array(clean_cols_aggregated[target_metric]))

        clean_wb.save(clean_file_name)
        noisy_wb.save(noisy_file_name)

    print('-- report generation has been completed --')
    print('\treport for clean setting is generated at ', clean_file_name)
    print('\treport for noisy setting is generated at ', noisy_file_name)


if __name__ == '__main__':
    main()
