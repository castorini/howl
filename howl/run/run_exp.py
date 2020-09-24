from .args import ArgumentParserBuilder, opt
import random
import os
import subprocess
import numpy as np
import torch
import sys
import time

from datetime import datetime
from openpyxl import Workbook, load_workbook
from tqdm import tqdm


def get_col(char, ind):
    return chr(ord(char) + ind)

def is_job_running(grep_command, count_command):
    out = subprocess.check_output('ps aux | grep ' + grep_command, shell=True)
    num_proc = out.decode('utf-8').count(count_command)
    return num_proc > 0

def run_batch_commands(commands, envs, grep_command = 'howl.run.train', count_command = 'python -m howl.run.train'):

    num_gpu = torch.cuda.device_count()
    print('availble GPU is', num_gpu)

    check_up_delay = 600 # 10mins
    counter = 0
    sleep_counter = 0

    for (command, env) in zip(commands, envs):
        for env_key, env_val in env.items():
            os.environ[env_key] = env_val


        if counter == num_gpu:
            sleep_counter = 0
            while is_job_running(grep_command, count_command):
                sleep_counter += 1
                print('job is running, wait time is {} min'.format(sleep_counter * check_up_delay / 60))
                sys.stdout.flush()
                sys.stderr.flush()
                time.sleep(check_up_delay)
            counter = 0

        new_env = os.environ.copy()
        new_env['CUDA_VISIBLE_DEVICES'] = str(counter)

        print('exec', command)
        proc = subprocess.Popen(command.split(), \
                                preexec_fn=os.setpgrp, \
                                env=new_env)

        print('execute processor {}'.format(proc.pid), command, env)
        time.sleep(60) # add some delay
        counter += 1


    sleep_counter = 0
    while is_job_running(grep_command, count_command):
        sleep_counter += 1
        print('job is running, wait time is {} min'.format(sleep_counter * check_up_delay / 60))
        sys.stdout.flush()
        sys.stderr.flush()
        time.sleep(check_up_delay)


def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--n',
                        type=int,
                        default=1,
                        help='number of experiments to run'),
                    opt('--hop_size',
                        type=float,
                        default=0.05,
                        help='hop size for threshold'),
                    opt('--dataset_path',
                        type=str,
                        default='/data/speaker-id-split-medium'),
                    opt('--exp_type', 
                        type=str, choices=['hey_ff', 'hey_snips'], default='hey_ff'),
                    opt('--noiseset_path',
                        type=str,
                        default='/data/MS-SNSD'))

    args = apb.parser.parse_args()

    random.seed()

    # Create clean workbook
    clean_wb = Workbook()
    noisy_wb = Workbook()
    now = datetime.now()

    thresholds = np.arange(0, 1.000001, args.hop_size)

    clean_sheets = {}
    noisy_sheets = {}


    col_map = {
        'Dev positive': 'B',
        'Dev noisy positive': 'B',
        'Dev negative': 'H',
        'Dev noisy negative': 'H',
        'Test positive': 'N',
        'Test noisy positive': 'N',
        'Test negative': 'T',
        'Test noisy negative': 'T'
    }

    total_map = {
        'Dev positive': '76',
        'Dev negative': '2531',
        'Test positive': '54',
        'Test negative': '2504'
    }
    metrics = ['threshold', 'tp', 'tn', 'fp', 'fn',]
    clean_col_names = ['Dev positive', 'Dev negative', 'Test positive', 'Test negative']

    clean_cols_aggregated = {}
    noisy_cols_aggregated = {}
    

    def fill_in_outline(sheet):

        sheet['A3'] = 'mean'
        sheet['A4'] = 'std'
        sheet['A5'] = 'p90'
        sheet['A6'] = 'p95'
        sheet['A7'] = 'p99'
        sheet['A8'] = 'sum'

        for col in clean_col_names:
            ind = col_map[col] + '1'
            sheet[ind] = col

            ind = get_col(col_map[col], 1) + '1'
            sheet[ind] = total_map[col]

            for i, metric in enumerate(metrics):
                ind = get_col(col_map[col], i) + '2'
                sheet[ind] = metric

    for idx, threshold in enumerate(thresholds):
        clean_sheets[threshold] = clean_wb.create_sheet(str(round(threshold, 2)), idx)
        fill_in_outline(clean_sheets[threshold])
        noisy_sheets[threshold] = noisy_wb.create_sheet(str(round(threshold, 2)), idx)
        fill_in_outline(noisy_sheets[threshold])

        for col in clean_col_names:
            for metric_idx, metric in enumerate(metrics):
                target_metric = str(round(threshold, 2)) + '_' + get_col(col_map[col], metric_idx)
                clean_cols_aggregated[target_metric] = []
                noisy_cols_aggregated[target_metric] = []


    os.system('mkdir -p exp_results')
    dt_string = now.strftime('%b-%d-%H-%M')

    clean_file_name = 'exp_results/'+args.exp_type+'_clean_' + dt_string + '.xlsx'
    noisy_file_name = 'exp_results/'+args.exp_type+'_noisy_' + dt_string + '.xlsx'

    clean_wb.save(clean_file_name)
    noisy_wb.save(noisy_file_name)
    print('\tclean exp result stored at ', clean_file_name)
    print('\tnoisy exp result stored at ', noisy_file_name)


    os.environ['DATASET_PATH'] = args.dataset_path
    os.environ['WEIGHT_DECAY'] = '0.00001'
    os.environ['NUM_EPOCHS'] = '300'
    os.environ['LEARNING_RATE'] = '0.01'
    os.environ['LR_DECAY'] = '0.98'
    os.environ['BATCH_SIZE'] = '16'
    os.environ['MAX_WINDOW_SIZE_SECONDS'] = '0.5'
    os.environ['USE_NOISE_DATASET'] = 'True'
    os.environ['NUM_MELS'] = '40'
    os.environ['INFERENCE_SEQUENCE'] = '[0,1,2]'

    os.environ['VOCAB'] = '[" hey","fire","fox"]'
    os.environ['NOISE_DATASET_PATH'] = args.noiseset_path


    def fill_aggregated(sheet, col_idx, results):
        sheet[col_idx + '3'] = str(results.mean())
        sheet[col_idx + '4'] = str(results.std())
        sheet[col_idx + '5'] = str(np.percentile(results, 90))
        sheet[col_idx + '6'] = str(np.percentile(results, 95))
        sheet[col_idx + '7'] = str(np.percentile(results, 99))
        sheet[col_idx + '8'] = str(results.sum())

    check_up_delay = 600 # 10mins

    seeds = []

    commands = []
    envs = []

    # train
    for i in range(args.n):
        seed = str(random.randint(1,1000000))
        seeds.append(seed)
        env = {}
        env['SEED'] = seed

        workspace_path = os.getcwd() + '/workspaces/exp_'+ args.exp_type +'_res8/' + str(seed)
        os.system('mkdir -p ' + workspace_path)

        command = 'python -m howl.run.train --model res8 --workspace ' + workspace_path + '  -i ' + args.dataset_path

        commands.append(command)
        envs.append(env)


    print('seeds: ', seeds)

    print('-- training --')

    run_batch_commands(commands, envs)

    commands = []
    envs = []

    for seed in seeds:
        for threshold_idx, threshold in enumerate(thresholds):

            env = {}
            env['SEED'] = seed
            env['INFERENCE_THRESHOLD'] = str(threshold)

            workspace_path = os.getcwd() + '/workspaces/exp_'+ args.exp_type +'_res8/' + str(seed)

            command = 'python -m howl.run.train --eval --model res8 --workspace ' + workspace_path + '  -i ' + args.dataset_path

            commands.append(command)
            envs.append(env)


    print('-- collecting metrics --')

    run_batch_commands(commands, envs)

    for i, seed in enumerate(seeds):
        for threshold_idx, threshold in enumerate(thresholds):

            clean_sheet = clean_sheets[threshold]
            noisy_sheet = noisy_sheets[threshold]

            row_index = str(i + 10)
            clean_sheet['A'+row_index] = seed
            noisy_sheet['A'+row_index] = seed

            workspace_path = os.getcwd() + '/workspaces/exp_'+ args.exp_type +'_res8/' + str(seed)

            log_path = workspace_path + '/' + str(round(threshold, 2)) + '_results.csv'
            raw_log = subprocess.check_output(['tail', '-n', '8', log_path]).decode('utf-8') 
            logs = raw_log.split('\n')

            for log in logs:
                if len(log) == 0:
                    break
                vals = log.split(',')
                key = vals[0]
                start_col = col_map[key]

                if 'noisy' in key:
                    for metric_idx, metric in enumerate(vals[1:]):
                        col_idx = get_col(start_col, metric_idx)
                        ind = col_idx + str(row_index)
                        noisy_sheet[ind] = metric

                        target_metric = str(round(threshold, 2)) + '_' + col_idx
                        noisy_cols_aggregated[target_metric].append(float(metric))

                        fill_aggregated(noisy_sheet, col_idx, np.array(noisy_cols_aggregated[target_metric]))
                else:
                    for metric_idx, metric in enumerate(vals[1:]):
                        col_idx = get_col(start_col, metric_idx)

                        ind = col_idx + str(row_index)
                        clean_sheet[ind] = metric

                        target_metric = str(round(threshold, 2)) + '_' + col_idx
                        clean_cols_aggregated[target_metric].append(float(metric))

                        fill_aggregated(clean_sheet, col_idx, np.array(clean_cols_aggregated[target_metric]))

        clean_wb.save(clean_file_name)
        noisy_wb.save(noisy_file_name)


if __name__ == '__main__':
    main()
