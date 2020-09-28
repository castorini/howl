from .args import ArgumentParserBuilder, opt
import random
import os
import subprocess
import numpy as np
import matplotlib.pyplot as plt

from datetime import datetime
from openpyxl import Workbook, load_workbook
from sklearn.metrics import roc_auc_score, roc_curve


def get_metrics(wb, thresholds):
    dev_far = []
    dev_frr = []

    test_far = []
    test_frr = []

    for threshold in thresholds:
        sheet = wb[str(threshold)]

        dev_tp = float(sheet['C3'].value)
        dev_fn = float(sheet['F3'].value)
        dev_tn = float(sheet['J3'].value)
        dev_fp = float(sheet['K3'].value)

        # dev_far.append(dev_fp / (dev_fp + dev_tn))
        dev_far.append(dev_fp / 3)
        dev_frr.append(dev_fn / (dev_fn + dev_tp))

        test_tp = float(sheet['O3'].value)
        test_fn = float(sheet['R3'].value)
        test_tn = float(sheet['V3'].value)
        test_fp = float(sheet['W3'].value)

        # test_far.append(test_fp / (test_fp + test_tn))
        test_far.append(test_fp / 3)
        test_frr.append(test_fn / (test_fn + test_tp))

    return dev_far, dev_frr, test_far, test_frr


def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--exp_timestemp',
                        type=str,
                        default="Sep-08-11-28"),
                    opt('--exp_type',
                        type=str,
                        default="hey_ff"))

    args = apb.parser.parse_args()

    print("clean file: ", "exp_results/"+args.exp_type"_clean_"+args.exp_timestemp+".xlsx")
    print("noisy file: ", "exp_results/"+args.exp_type"_noisy_"+args.exp_timestemp+".xlsx")
    clean_wb = load_workbook("exp_results/"+args.exp_type"_clean_"+args.exp_timestemp+".xlsx")
    noisy_wb = load_workbook("exp_results/"+args.exp_type"_noisy_"+args.exp_timestemp+".xlsx")

    thresholds = []

    for name in clean_wb.sheetnames:
        try:
            thresholds.append(float(name))
        except ValueError:
            print("Not a float: ", name)

    clean_dev_far, clean_dev_frr, clean_test_far, clean_test_frr = get_metrics(clean_wb, thresholds)
    noisy_dev_far, noisy_dev_frr, noisy_test_far, noisy_test_frr  = get_metrics(noisy_wb, thresholds)


    plt.title('ROC curve')
    plt.xlabel('False Acceptance Per Hour')
    plt.ylabel('False Rejection Rate')

    print("thresholds:", thresholds)
    print("clean_dev_faph:", [round(num, 3) for num in clean_dev_far])
    print("clean_dev_frr:", [round(num, 3) for num in clean_dev_frr])
    print("clean_test_faph:", [round(num, 3) for num in clean_test_far])
    print("clean_test_frr:", [round(num, 3) for num in clean_test_frr])

    print("noisy_dev_faph:", [round(num, 3) for num in noisy_dev_far])
    print("noisy_dev_frr:", [round(num, 3) for num in noisy_dev_frr])
    print("noisy_test_faph:", [round(num, 3) for num in noisy_test_far])
    print("noisy_test_frr:", [round(num, 3) for num in noisy_test_frr])

    plt.plot(clean_dev_far, clean_dev_frr, 'r--+', label = 'clean dev')
    plt.plot(clean_test_far, clean_test_frr, 'r-+', label = 'clean test')
    plt.plot(noisy_dev_far, noisy_dev_frr, 'b--+', label = 'noisy dev')
    plt.plot(noisy_test_far, noisy_test_frr, 'b-+', label = 'noisy test')

    plt.legend()

    # plt.show()
    plt.savefig("exp_results/" args.exp_type+ "_" +args.exp_timestemp + '.png')

if __name__ == '__main__':
    main()
