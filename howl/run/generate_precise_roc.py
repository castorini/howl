from .args import ArgumentParserBuilder, opt
import random
import os
import subprocess
import numpy as np
import matplotlib.pyplot as plt

from datetime import datetime
from openpyxl import Workbook, load_workbook
from sklearn.metrics import roc_auc_score, roc_curve


def get_metrics(exp_type, wb, thresholds):
    dev_far = []
    dev_frr = []

    test_far = []
    test_frr = []
    if exp_type == "hey_ff":
        total_train_len = 111928.04543750011
        total_dev_len = 10679.505062500015
        total_test_len = 10364.291000000001
        dev_pos_audio_len = 775.9072499999995
        dev_neg_audio_len = 9903.597812500018
        test_pos_audio_len = 874.6844999999998
        test_neg_audio_len = 9489.606499999993

    elif exp_type == "hey_snips":
        # to be updated
        total_dev_len = 46066.6921250002
        total_test_len = 47047.301562499844
    

    for threshold in thresholds:
        sheet = wb[str(threshold)]

        dev_tp = float(sheet['C3'].value)
        dev_fn = float(sheet['F3'].value)
        dev_tn = float(sheet['J3'].value)
        dev_fp = float(sheet['K3'].value)

        # dev_far.append(dev_fp / (dev_fp + dev_tn))
        dev_far.append(dev_fp / (total_dev_len / 3600))
        dev_frr.append(dev_fn / (dev_fn + dev_tp))

        test_tp = float(sheet['O3'].value)
        test_fn = float(sheet['R3'].value)
        test_tn = float(sheet['V3'].value)
        test_fp = float(sheet['W3'].value)

        # test_far.append(test_fp / (test_fp + test_tn))
        test_far.append(test_fp / (total_test_len / 3600))
        test_frr.append(test_fn / (test_fn + test_tp))

    return dev_far, dev_frr, test_far, test_frr


def main():
    apb = ArgumentParserBuilder()
    apb.add_options(opt('--exp_timestemp',
                        type=str,
                        default="Sep-08-11-28"),
                    opt('--exp_type',
                        type=str,
                        default="hey_ff"))

    args = apb.parser.parse_args()

    print("exp type: ", args.exp_type)

    print("clean file: ", "exp_results/"+args.exp_type+"_clean_"+args.exp_timestemp+".xlsx")
    print("noisy file: ", "exp_results/"+args.exp_type+"_noisy_"+args.exp_timestemp+".xlsx")
    clean_wb = load_workbook("exp_results/hey_snips_v2/"+args.exp_type+"_clean_"+args.exp_timestemp+".xlsx")
    noisy_wb = load_workbook("exp_results/hey_snips_v2/"+args.exp_type+"_noisy_"+args.exp_timestemp+".xlsx")
    precise_wb = load_workbook("exp_results/hey_snips_v2/"+args.exp_type+"_precise_Oct-13-08-31.xlsx")


    thresholds = []

    for name in clean_wb.sheetnames:
        try:
            thresholds.append(float(name))
        except ValueError:
            print("Not a float: ", name)

    clean_dev_far, clean_dev_frr, clean_test_far, clean_test_frr = get_metrics(args.exp_type, clean_wb, thresholds)
    noisy_dev_far, noisy_dev_frr, noisy_test_far, noisy_test_frr  = get_metrics(args.exp_type, noisy_wb, thresholds)

    precise_dev_far, precise_dev_frr, precise_test_far, precise_test_frr  = get_metrics(args.exp_type, precise_wb, thresholds)


    plt.rcParams.update({'font.size': 12})

    # plt.title('ROC curve')
    plt.xlabel('False Alarms Per Hour')
    plt.ylabel('False Rejection Rate')

    print("thresholds:", thresholds)
    print("clean_dev_faph:", [round(num, 3) for num in clean_dev_far])
    print("clean_dev_frr:", [round(num, 3) for num in clean_dev_frr])
    print("clean_test_faph:", [round(num, 3) for num in clean_test_far])
    print("clean_test_frr:", [round(num, 3) for num in clean_test_frr])

    print("noisy_dev_faph:", [round(num, 3) for num in noisy_dev_far])
    print("noisy_dev_frr:", [round(num, 3) for num in noisy_dev_frr])
    print("noisy_test_faph:", [round(num, 3) for num in noisy_test_far])
    print("noisy_test_frr:", [round(num, 3) for num in noisy_test_frr])

    plt.plot(clean_dev_far[1:-1], clean_dev_frr[1:-1], '--+', color='tab:blue', label = 'clean dev')
    plt.plot(clean_test_far[1:-1], clean_test_frr[1:-1], '-+', color='tab:blue', label = 'clean test')
    plt.plot(noisy_dev_far[1:-1], noisy_dev_frr[1:-1], '--+', color='tab:orange', label = 'noisy dev')
    plt.plot(noisy_test_far[1:-1], noisy_test_frr[1:-1], '-+', color='tab:orange', label = 'noisy test')
    plt.plot(precise_dev_far[1:-1], precise_dev_frr[1:-1], '--+', color='tab:green', label = 'precise dev')
    plt.plot(precise_test_far[1:-1], precise_test_frr[1:-1], '-+', color='tab:green', label = 'precise test')

    plt.grid()

    plt.legend()

    # plt.show()
    plt.savefig("exp_results/" + args.exp_type+ "_" +args.exp_timestemp + '.pdf')

if __name__ == '__main__':
    main()
